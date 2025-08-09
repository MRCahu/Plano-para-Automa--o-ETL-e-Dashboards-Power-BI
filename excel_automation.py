"""
Script de Automação para Excel
Demonstra integração Python-Excel e geração de código VBA
Autor: Sistema de Automação de Dados
Data: 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class ExcelAutomation:
    """Classe para automação de Excel"""
    
    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file
        self.workbook = None
        self.data = None
    
    def load_data(self):
        """Carrega dados do arquivo Excel"""
        try:
            self.data = pd.read_excel(self.input_file, sheet_name='Dados_Principais')
            print(f"Dados carregados: {len(self.data)} registros")
            return True
        except Exception as e:
            print(f"Erro ao carregar dados: {e}")
            return False
    
    def create_formatted_workbook(self):
        """Cria planilha formatada com estilos"""
        try:
            # Criar novo workbook
            self.workbook = openpyxl.Workbook()
            
            # Remover planilha padrão
            self.workbook.remove(self.workbook.active)
            
            # Criar planilhas
            self._create_dashboard_sheet()
            self._create_data_sheet()
            self._create_summary_sheet()
            self._create_charts_sheet()
            
            print("Workbook formatado criado com sucesso")
            return True
            
        except Exception as e:
            print(f"Erro ao criar workbook: {e}")
            return False
    
    def _create_dashboard_sheet(self):
        """Cria planilha de dashboard"""
        ws = self.workbook.create_sheet("Dashboard", 0)
        
        # Título principal
        ws['A1'] = "DASHBOARD FINANCEIRO"
        ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:F1')
        
        # KPIs principais
        kpis = [
            ("Total de Transações", len(self.data)),
            ("Valor Total", f"R$ {self.data['Valor'].sum():,.2f}"),
            ("Valor Médio", f"R$ {self.data['Valor'].mean():,.2f}"),
            ("Departamentos", self.data['Departamento'].nunique())
        ]
        
        row = 3
        for i, (label, value) in enumerate(kpis):
            col = chr(65 + i * 2)  # A, C, E, G
            
            # Label
            ws[f'{col}{row}'] = label
            ws[f'{col}{row}'].font = Font(bold=True, size=12)
            ws[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Valor
            ws[f'{col}{row+1}'] = str(value)
            ws[f'{col}{row+1}'].font = Font(size=14, bold=True, color="366092")
            ws[f'{col}{row+1}'].alignment = Alignment(horizontal="center")
            
            # Merge cells para KPI
            ws.merge_cells(f'{col}{row}:{chr(ord(col)+1)}{row}')
            ws.merge_cells(f'{col}{row+1}:{chr(ord(col)+1)}{row+1}')
        
        # Ajustar largura das colunas
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 25
    
    def _create_data_sheet(self):
        """Cria planilha com dados formatados"""
        ws = self.workbook.create_sheet("Dados")
        
        # Adicionar dados
        for r in dataframe_to_rows(self.data, index=False, header=True):
            ws.append(r)
        
        # Formatação do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatação das células de dados
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Não aplicar a células do cabeçalho
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self):
        """Cria planilha de resumo"""
        ws = self.workbook.create_sheet("Resumo")
        
        # Título
        ws['A1'] = "RESUMO POR DEPARTAMENTO"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells('A1:E1')
        
        # Criar resumo por departamento
        summary = self.data.groupby('Departamento').agg({
            'Valor': ['count', 'sum', 'mean', 'max', 'min']
        }).round(2)
        
        summary.columns = ['Qtd_Transações', 'Total_Valor', 'Média_Valor', 'Maior_Valor', 'Menor_Valor']
        summary = summary.reset_index()
        
        # Adicionar cabeçalhos
        headers = ['Departamento', 'Qtd Transações', 'Total Valor', 'Média Valor', 'Maior Valor', 'Menor Valor']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Adicionar dados do resumo
        for i, row in summary.iterrows():
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i+4, column=j, value=value)
                if j > 2:  # Formatar valores monetários
                    if isinstance(value, (int, float)):
                        cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="center")
        
        # Formatação
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar largura das colunas
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18
    
    def _create_charts_sheet(self):
        """Cria planilha com gráficos"""
        ws = self.workbook.create_sheet("Gráficos")
        
        # Preparar dados para gráficos
        dept_summary = self.data.groupby('Departamento')['Valor'].sum().reset_index()
        
        # Adicionar dados para gráfico
        ws['A1'] = "Departamento"
        ws['B1'] = "Total Valor"
        
        for i, (dept, valor) in enumerate(zip(dept_summary['Departamento'], dept_summary['Valor']), 2):
            ws[f'A{i}'] = dept
            ws[f'B{i}'] = valor
        
        # Criar gráfico de barras
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Valor Total por Departamento"
        chart.y_axis.title = 'Valor (R$)'
        chart.x_axis.title = 'Departamento'
        
        data = Reference(ws, min_col=2, min_row=1, max_row=len(dept_summary)+1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(dept_summary)+1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 15
        chart.width = 20
        
        ws.add_chart(chart, "D2")
    
    def generate_vba_code(self):
        """Gera código VBA para automação"""
        vba_code = '''
Sub AutomatizarRelatorio()
    '
    ' Macro para automatizar geração de relatórios
    ' Criada pelo Sistema de Automação de Dados
    '
    
    Dim ws As Worksheet
    Dim lastRow As Long
    Dim dataRange As Range
    
    ' Definir planilha de trabalho
    Set ws = ActiveSheet
    
    ' Encontrar última linha com dados
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
    
    ' Definir intervalo de dados
    Set dataRange = ws.Range("A1:Z" & lastRow)
    
    ' Aplicar formatação automática
    Call FormatarDados(dataRange)
    
    ' Criar tabela dinâmica
    Call CriarTabelaDinamica(dataRange)
    
    ' Gerar gráficos
    Call GerarGraficos(dataRange)
    
    MsgBox "Relatório automatizado gerado com sucesso!", vbInformation
    
End Sub

Sub FormatarDados(dataRange As Range)
    '
    ' Formatar dados automaticamente
    '
    
    With dataRange
        .Font.Name = "Calibri"
        .Font.Size = 11
        .Borders.LineStyle = xlContinuous
        .Borders.Weight = xlThin
    End With
    
    ' Formatação do cabeçalho
    With dataRange.Rows(1)
        .Font.Bold = True
        .Interior.Color = RGB(54, 96, 146)
        .Font.Color = RGB(255, 255, 255)
        .HorizontalAlignment = xlCenter
    End With
    
    ' Auto-ajustar colunas
    dataRange.Columns.AutoFit
    
End Sub

Sub CriarTabelaDinamica(dataRange As Range)
    '
    ' Criar tabela dinâmica automaticamente
    '
    
    Dim pt As PivotTable
    Dim pc As PivotCache
    Dim newWs As Worksheet
    
    ' Criar nova planilha para tabela dinâmica
    Set newWs = Worksheets.Add
    newWs.Name = "Tabela_Dinamica_" & Format(Now, "hhmmss")
    
    ' Criar cache da tabela dinâmica
    Set pc = ActiveWorkbook.PivotCaches.Create( _
        SourceType:=xlDatabase, _
        SourceData:=dataRange)
    
    ' Criar tabela dinâmica
    Set pt = pc.CreatePivotTable( _
        TableDestination:=newWs.Range("A1"), _
        TableName:="TabelaDinamica1")
    
    ' Configurar campos da tabela dinâmica
    With pt
        .PivotFields("Departamento").Orientation = xlRowField
        .PivotFields("Categoria").Orientation = xlRowField
        .PivotFields("Valor").Orientation = xlDataField
        .PivotFields("Sum of Valor").NumberFormat = "R$ #,##0.00"
    End With
    
End Sub

Sub GerarGraficos(dataRange As Range)
    '
    ' Gerar gráficos automaticamente
    '
    
    Dim chartObj As ChartObject
    Dim newWs As Worksheet
    
    ' Criar nova planilha para gráficos
    Set newWs = Worksheets.Add
    newWs.Name = "Graficos_" & Format(Now, "hhmmss")
    
    ' Criar gráfico de colunas
    Set chartObj = newWs.ChartObjects.Add(50, 50, 400, 300)
    
    With chartObj.Chart
        .SetSourceData dataRange
        .ChartType = xlColumnClustered
        .HasTitle = True
        .ChartTitle.Text = "Análise de Dados - " & Format(Now, "dd/mm/yyyy")
        .Axes(xlCategory, xlPrimary).HasTitle = True
        .Axes(xlCategory, xlPrimary).AxisTitle.Text = "Categorias"
        .Axes(xlValue, xlPrimary).HasTitle = True
        .Axes(xlValue, xlPrimary).AxisTitle.Text = "Valores"
    End With
    
End Sub

Sub ExportarParaPDF()
    '
    ' Exportar planilha para PDF
    '
    
    Dim fileName As String
    fileName = ThisWorkbook.Path & "\\Relatorio_" & Format(Now, "yyyy-mm-dd_hhmmss") & ".pdf"
    
    ActiveSheet.ExportAsFixedFormat _
        Type:=xlTypePDF, _
        fileName:=fileName, _
        Quality:=xlQualityStandard, _
        IncludeDocProps:=True, _
        IgnorePrintAreas:=False, _
        OpenAfterPublish:=True
    
    MsgBox "Relatório exportado para: " & fileName, vbInformation
    
End Sub

Sub AtualizarDadosAutomaticamente()
    '
    ' Atualizar dados de fonte externa
    '
    
    Dim conn As WorkbookConnection
    
    ' Atualizar todas as conexões
    For Each conn In ThisWorkbook.Connections
        conn.Refresh
    Next conn
    
    ' Atualizar tabelas dinâmicas
    Dim pt As PivotTable
    For Each pt In ActiveSheet.PivotTables
        pt.RefreshTable
    Next pt
    
    MsgBox "Dados atualizados com sucesso!", vbInformation
    
End Sub
'''
        
        # Salvar código VBA em arquivo
        with open('/home/ubuntu/codigo_vba_automacao.txt', 'w', encoding='utf-8') as f:
            f.write(vba_code)
        
        print("Código VBA gerado e salvo em: codigo_vba_automacao.txt")
        return vba_code
    
    def save_workbook(self):
        """Salva o workbook"""
        try:
            self.workbook.save(self.output_file)
            print(f"Arquivo Excel salvo: {self.output_file}")
            return True
        except Exception as e:
            print(f"Erro ao salvar arquivo: {e}")
            return False
    
    def run_automation(self):
        """Executa automação completa"""
        print("=== INICIANDO AUTOMAÇÃO EXCEL ===")
        
        if not self.load_data():
            return False
        
        if not self.create_formatted_workbook():
            return False
        
        self.generate_vba_code()
        
        if not self.save_workbook():
            return False
        
        print("=== AUTOMAÇÃO EXCEL CONCLUÍDA ===")
        return True

# Exemplo de uso
if __name__ == "__main__":
    # Verificar se arquivo de entrada existe
    input_file = "dados_ficticios_1000_linhas.xlsx"
    output_file = "relatorio_automatizado.xlsx"
    
    if os.path.exists(input_file):
        automation = ExcelAutomation(input_file, output_file)
        automation.run_automation()
    else:
        print(f"Arquivo de entrada não encontrado: {input_file}")
        print("Execute primeiro o script de geração de dados fictícios.")

