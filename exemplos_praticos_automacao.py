"""
Exemplos Práticos de Automação de Dados
Complemento ao Guia Completo de Automação Excel e Power BI
Autor: Manus AI
Data: 2025
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import json
import logging

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class RelatorioFinanceiroAutomatizado:
    """
    Classe para geração automatizada de relatórios financeiros
    Demonstra integração completa de ETL, formatação e distribuição
    """
    
    def __init__(self, dados_origem, config_email=None):
        self.dados_origem = dados_origem
        self.config_email = config_email
        self.dados_processados = None
        self.relatorio_excel = None
        
    def extrair_dados(self):
        """Extrai dados de múltiplas fontes"""
        try:
            # Simular extração de diferentes fontes
            if isinstance(self.dados_origem, str) and self.dados_origem.endswith('.xlsx'):
                self.dados_brutos = pd.read_excel(self.dados_origem)
            elif isinstance(self.dados_origem, str) and self.dados_origem.endswith('.csv'):
                self.dados_brutos = pd.read_csv(self.dados_origem)
            else:
                self.dados_brutos = self.dados_origem
            
            logging.info(f"Dados extraídos: {len(self.dados_brutos)} registros")
            return True
            
        except Exception as e:
            logging.error(f"Erro na extração: {e}")
            return False
    
    def transformar_dados(self):
        """Aplica transformações específicas para relatório financeiro"""
        try:
            df = self.dados_brutos.copy()
            
            # Converter data se necessário
            if 'Data' in df.columns:
                df['Data'] = pd.to_datetime(df['Data'])
                df['Mes'] = df['Data'].dt.month
                df['Ano'] = df['Data'].dt.year
                df['Trimestre'] = df['Data'].dt.quarter
            
            # Categorizar valores
            if 'Valor' in df.columns:
                df['Faixa_Valor'] = pd.cut(df['Valor'], 
                                         bins=[0, 1000, 5000, 15000, float('inf')],
                                         labels=['Baixo', 'Médio', 'Alto', 'Muito Alto'])
            
            # Calcular métricas financeiras
            df['Valor_Acumulado'] = df.groupby('Departamento')['Valor'].cumsum()
            df['Percentual_Total'] = (df['Valor'] / df['Valor'].sum()) * 100
            
            # Identificar outliers
            Q1 = df['Valor'].quantile(0.25)
            Q3 = df['Valor'].quantile(0.75)
            IQR = Q3 - Q1
            df['Outlier'] = (df['Valor'] < (Q1 - 1.5 * IQR)) | (df['Valor'] > (Q3 + 1.5 * IQR))
            
            self.dados_processados = df
            logging.info("Transformações aplicadas com sucesso")
            return True
            
        except Exception as e:
            logging.error(f"Erro na transformação: {e}")
            return False
    
    def gerar_relatorio_excel(self, nome_arquivo):
        """Gera relatório Excel formatado com múltiplas abas"""
        try:
            with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
                # Aba principal com dados
                self.dados_processados.to_excel(writer, sheet_name='Dados_Detalhados', index=False)
                
                # Aba de resumo executivo
                resumo = self._criar_resumo_executivo()
                resumo.to_excel(writer, sheet_name='Resumo_Executivo', index=False)
                
                # Aba de análise por departamento
                analise_dept = self._criar_analise_departamento()
                analise_dept.to_excel(writer, sheet_name='Analise_Departamento', index=False)
                
                # Aba de tendências mensais
                tendencias = self._criar_analise_temporal()
                tendencias.to_excel(writer, sheet_name='Tendencias_Mensais', index=False)
            
            # Aplicar formatação avançada
            self._aplicar_formatacao_avancada(nome_arquivo)
            
            self.relatorio_excel = nome_arquivo
            logging.info(f"Relatório Excel gerado: {nome_arquivo}")
            return True
            
        except Exception as e:
            logging.error(f"Erro na geração do Excel: {e}")
            return False
    
    def _criar_resumo_executivo(self):
        """Cria resumo executivo com KPIs principais"""
        resumo_data = {
            'Métrica': [
                'Total de Transações',
                'Valor Total (R$)',
                'Valor Médio (R$)',
                'Maior Transação (R$)',
                'Menor Transação (R$)',
                'Departamentos Ativos',
                'Período Analisado',
                'Outliers Identificados',
                'Concentração Top 3 Departamentos (%)'
            ],
            'Valor': [
                len(self.dados_processados),
                f"{self.dados_processados['Valor'].sum():,.2f}",
                f"{self.dados_processados['Valor'].mean():,.2f}",
                f"{self.dados_processados['Valor'].max():,.2f}",
                f"{self.dados_processados['Valor'].min():,.2f}",
                self.dados_processados['Departamento'].nunique(),
                f"{self.dados_processados['Data'].min().strftime('%d/%m/%Y')} a {self.dados_processados['Data'].max().strftime('%d/%m/%Y')}",
                self.dados_processados['Outlier'].sum(),
                f"{self.dados_processados.groupby('Departamento')['Valor'].sum().nlargest(3).sum() / self.dados_processados['Valor'].sum() * 100:.1f}%"
            ]
        }
        return pd.DataFrame(resumo_data)
    
    def _criar_analise_departamento(self):
        """Cria análise detalhada por departamento"""
        analise = self.dados_processados.groupby('Departamento').agg({
            'Valor': ['count', 'sum', 'mean', 'std', 'min', 'max'],
            'Outlier': 'sum'
        }).round(2)
        
        analise.columns = ['Qtd_Transacoes', 'Total_Valor', 'Media_Valor', 
                          'Desvio_Padrao', 'Min_Valor', 'Max_Valor', 'Outliers']
        
        # Adicionar percentual do total
        analise['Percentual_Total'] = (analise['Total_Valor'] / analise['Total_Valor'].sum() * 100).round(2)
        
        # Adicionar ranking
        analise['Ranking_Valor'] = analise['Total_Valor'].rank(method='dense', ascending=False).astype(int)
        
        return analise.reset_index()
    
    def _criar_analise_temporal(self):
        """Cria análise de tendências temporais"""
        temporal = self.dados_processados.groupby(['Ano', 'Mes']).agg({
            'Valor': ['count', 'sum', 'mean'],
            'Departamento': 'nunique'
        }).round(2)
        
        temporal.columns = ['Qtd_Transacoes', 'Total_Valor', 'Media_Valor', 'Departamentos_Ativos']
        temporal = temporal.reset_index()
        
        # Calcular crescimento mês a mês
        temporal['Crescimento_MoM'] = temporal['Total_Valor'].pct_change() * 100
        
        # Calcular média móvel de 3 meses
        temporal['Media_Movel_3M'] = temporal['Total_Valor'].rolling(window=3).mean()
        
        return temporal
    
    def _aplicar_formatacao_avancada(self, nome_arquivo):
        """Aplica formatação avançada ao arquivo Excel"""
        wb = openpyxl.load_workbook(nome_arquivo)
        
        # Formatação para cada aba
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formatação do cabeçalho
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Auto-ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(nome_arquivo)
    
    def gerar_graficos_analise(self, diretorio_saida):
        """Gera gráficos de análise e salva como imagens"""
        try:
            # Configurar estilo
            plt.style.use('seaborn-v0_8')
            fig_size = (12, 8)
            
            # Gráfico 1: Distribuição por Departamento
            plt.figure(figsize=fig_size)
            dept_valores = self.dados_processados.groupby('Departamento')['Valor'].sum().sort_values(ascending=True)
            dept_valores.plot(kind='barh')
            plt.title('Distribuição de Valores por Departamento', fontsize=16, fontweight='bold')
            plt.xlabel('Valor Total (R$)')
            plt.ylabel('Departamento')
            plt.tight_layout()
            plt.savefig(f'{diretorio_saida}/distribuicao_departamentos.png', dpi=300, bbox_inches='tight')
            plt.close()
            
            # Gráfico 2: Tendência Temporal
            plt.figure(figsize=fig_size)
            temporal = self.dados_processados.groupby(self.dados_processados['Data'].dt.to_period('M'))['Valor'].sum()
            temporal.plot(kind='line', marker='o')
            plt.title('Tendência Temporal de Valores', fontsize=16, fontweight='bold')
            plt.xlabel('Período')
            plt.ylabel('Valor Total (R$)')
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(f'{diretorio_saida}/tendencia_temporal.png', dpi=300, bbox_inches='tight')
            plt.close()
            
            # Gráfico 3: Distribuição de Valores (Histograma)
            plt.figure(figsize=fig_size)
            plt.hist(self.dados_processados['Valor'], bins=30, alpha=0.7, edgecolor='black')
            plt.title('Distribuição de Valores das Transações', fontsize=16, fontweight='bold')
            plt.xlabel('Valor (R$)')
            plt.ylabel('Frequência')
            plt.tight_layout()
            plt.savefig(f'{diretorio_saida}/distribuicao_valores.png', dpi=300, bbox_inches='tight')
            plt.close()
            
            # Gráfico 4: Top 10 Transações
            plt.figure(figsize=fig_size)
            top_10 = self.dados_processados.nlargest(10, 'Valor')
            plt.barh(range(len(top_10)), top_10['Valor'])
            plt.yticks(range(len(top_10)), [f"{row['Departamento'][:15]}..." for _, row in top_10.iterrows()])
            plt.title('Top 10 Maiores Transações', fontsize=16, fontweight='bold')
            plt.xlabel('Valor (R$)')
            plt.tight_layout()
            plt.savefig(f'{diretorio_saida}/top_10_transacoes.png', dpi=300, bbox_inches='tight')
            plt.close()
            
            logging.info(f"Gráficos salvos em: {diretorio_saida}")
            return True
            
        except Exception as e:
            logging.error(f"Erro na geração de gráficos: {e}")
            return False
    
    def enviar_relatorio_email(self, destinatarios, assunto=None):
        """Envia relatório por email automaticamente"""
        if not self.config_email:
            logging.error("Configuração de email não fornecida")
            return False
        
        try:
            # Configurar mensagem
            msg = MIMEMultipart()
            msg['From'] = self.config_email['remetente']
            msg['To'] = ', '.join(destinatarios)
            msg['Subject'] = assunto or f"Relatório Financeiro Automatizado - {datetime.now().strftime('%d/%m/%Y')}"
            
            # Corpo do email
            corpo_email = f"""
            Prezados,
            
            Segue em anexo o relatório financeiro automatizado gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}.
            
            Resumo dos dados processados:
            - Total de transações: {len(self.dados_processados):,}
            - Valor total: R$ {self.dados_processados['Valor'].sum():,.2f}
            - Período: {self.dados_processados['Data'].min().strftime('%d/%m/%Y')} a {self.dados_processados['Data'].max().strftime('%d/%m/%Y')}
            
            Este relatório foi gerado automaticamente pelo sistema de automação de dados.
            
            Atenciosamente,
            Sistema de Automação Financeira
            """
            
            msg.attach(MIMEText(corpo_email, 'plain'))
            
            # Anexar arquivo Excel
            if self.relatorio_excel and os.path.exists(self.relatorio_excel):
                with open(self.relatorio_excel, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {os.path.basename(self.relatorio_excel)}'
                    )
                    msg.attach(part)
            
            # Enviar email
            server = smtplib.SMTP(self.config_email['servidor'], self.config_email['porta'])
            server.starttls()
            server.login(self.config_email['usuario'], self.config_email['senha'])
            server.send_message(msg)
            server.quit()
            
            logging.info(f"Relatório enviado para: {', '.join(destinatarios)}")
            return True
            
        except Exception as e:
            logging.error(f"Erro no envio de email: {e}")
            return False
    
    def executar_pipeline_completo(self, nome_arquivo_saida, destinatarios_email=None):
        """Executa pipeline completo de geração de relatório"""
        logging.info("=== INICIANDO PIPELINE DE RELATÓRIO FINANCEIRO ===")
        
        # Extração
        if not self.extrair_dados():
            return False
        
        # Transformação
        if not self.transformar_dados():
            return False
        
        # Geração do relatório
        if not self.gerar_relatorio_excel(nome_arquivo_saida):
            return False
        
        # Geração de gráficos
        diretorio_graficos = os.path.dirname(nome_arquivo_saida) or '.'
        self.gerar_graficos_analise(diretorio_graficos)
        
        # Envio por email (se configurado)
        if destinatarios_email and self.config_email:
            self.enviar_relatorio_email(destinatarios_email)
        
        logging.info("=== PIPELINE CONCLUÍDO COM SUCESSO ===")
        return True

class MonitorKPIs:
    """
    Classe para monitoramento de KPIs em tempo real
    Demonstra implementação de alertas e dashboards automatizados
    """
    
    def __init__(self, dados, thresholds):
        self.dados = dados
        self.thresholds = thresholds
        self.alertas = []
    
    def calcular_kpis(self):
        """Calcula KPIs principais"""
        kpis = {
            'total_transacoes': len(self.dados),
            'valor_total': self.dados['Valor'].sum(),
            'valor_medio': self.dados['Valor'].mean(),
            'transacoes_hoje': len(self.dados[self.dados['Data'].dt.date == datetime.now().date()]),
            'departamentos_ativos': self.dados['Departamento'].nunique(),
            'outliers_detectados': self.dados['Outlier'].sum() if 'Outlier' in self.dados.columns else 0
        }
        return kpis
    
    def verificar_alertas(self, kpis):
        """Verifica se KPIs estão dentro dos thresholds"""
        self.alertas = []
        
        for kpi, valor in kpis.items():
            if kpi in self.thresholds:
                threshold = self.thresholds[kpi]
                
                if 'min' in threshold and valor < threshold['min']:
                    self.alertas.append({
                        'kpi': kpi,
                        'valor': valor,
                        'threshold': threshold['min'],
                        'tipo': 'BAIXO',
                        'timestamp': datetime.now()
                    })
                
                if 'max' in threshold and valor > threshold['max']:
                    self.alertas.append({
                        'kpi': kpi,
                        'valor': valor,
                        'threshold': threshold['max'],
                        'tipo': 'ALTO',
                        'timestamp': datetime.now()
                    })
        
        return self.alertas
    
    def gerar_dashboard_kpis(self, nome_arquivo):
        """Gera dashboard visual dos KPIs"""
        kpis = self.calcular_kpis()
        
        fig, axes = plt.subplots(2, 3, figsize=(15, 10))
        fig.suptitle('Dashboard de KPIs - Monitoramento em Tempo Real', fontsize=16, fontweight='bold')
        
        # KPI 1: Total de Transações
        axes[0, 0].text(0.5, 0.5, f"{kpis['total_transacoes']:,}", 
                       ha='center', va='center', fontsize=24, fontweight='bold')
        axes[0, 0].set_title('Total de Transações')
        axes[0, 0].axis('off')
        
        # KPI 2: Valor Total
        axes[0, 1].text(0.5, 0.5, f"R$ {kpis['valor_total']:,.0f}", 
                       ha='center', va='center', fontsize=20, fontweight='bold')
        axes[0, 1].set_title('Valor Total')
        axes[0, 1].axis('off')
        
        # KPI 3: Valor Médio
        axes[0, 2].text(0.5, 0.5, f"R$ {kpis['valor_medio']:,.0f}", 
                       ha='center', va='center', fontsize=20, fontweight='bold')
        axes[0, 2].set_title('Valor Médio')
        axes[0, 2].axis('off')
        
        # KPI 4: Transações Hoje
        axes[1, 0].text(0.5, 0.5, f"{kpis['transacoes_hoje']:,}", 
                       ha='center', va='center', fontsize=24, fontweight='bold')
        axes[1, 0].set_title('Transações Hoje')
        axes[1, 0].axis('off')
        
        # KPI 5: Departamentos Ativos
        axes[1, 1].text(0.5, 0.5, f"{kpis['departamentos_ativos']}", 
                       ha='center', va='center', fontsize=24, fontweight='bold')
        axes[1, 1].set_title('Departamentos Ativos')
        axes[1, 1].axis('off')
        
        # KPI 6: Outliers
        cor_outlier = 'red' if kpis['outliers_detectados'] > 0 else 'green'
        axes[1, 2].text(0.5, 0.5, f"{kpis['outliers_detectados']}", 
                       ha='center', va='center', fontsize=24, fontweight='bold', color=cor_outlier)
        axes[1, 2].set_title('Outliers Detectados')
        axes[1, 2].axis('off')
        
        plt.tight_layout()
        plt.savefig(nome_arquivo, dpi=300, bbox_inches='tight')
        plt.close()
        
        return kpis

def exemplo_uso_completo():
    """
    Exemplo de uso completo das classes de automação
    Demonstra pipeline end-to-end
    """
    
    # Configuração de email (exemplo)
    config_email = {
        'servidor': 'smtp.gmail.com',
        'porta': 587,
        'usuario': 'seu_email@gmail.com',
        'senha': 'sua_senha_app',
        'remetente': 'seu_email@gmail.com'
    }
    
    # Thresholds para monitoramento
    thresholds = {
        'valor_total': {'min': 100000, 'max': 1000000},
        'transacoes_hoje': {'min': 10, 'max': 100},
        'outliers_detectados': {'max': 5}
    }
    
    try:
        # Carregar dados (assumindo que o arquivo existe)
        arquivo_dados = 'dados_ficticios_1000_linhas.xlsx'
        
        if os.path.exists(arquivo_dados):
            # Gerar relatório financeiro automatizado
            relatorio = RelatorioFinanceiroAutomatizado(arquivo_dados, config_email)
            sucesso = relatorio.executar_pipeline_completo(
                'relatorio_financeiro_automatizado.xlsx',
                # destinatarios_email=['gestor@empresa.com', 'financeiro@empresa.com']
            )
            
            if sucesso:
                print("✅ Relatório financeiro gerado com sucesso!")
                
                # Monitoramento de KPIs
                dados = pd.read_excel(arquivo_dados)
                monitor = MonitorKPIs(dados, thresholds)
                kpis = monitor.gerar_dashboard_kpis('dashboard_kpis.png')
                alertas = monitor.verificar_alertas(kpis)
                
                if alertas:
                    print("⚠️  Alertas detectados:")
                    for alerta in alertas:
                        print(f"   - {alerta['kpi']}: {alerta['valor']} ({alerta['tipo']})")
                else:
                    print("✅ Todos os KPIs dentro dos parâmetros normais")
                
                print(f"📊 Dashboard de KPIs salvo em: dashboard_kpis.png")
            
        else:
            print(f"❌ Arquivo de dados não encontrado: {arquivo_dados}")
            print("Execute primeiro o script de geração de dados fictícios.")
    
    except Exception as e:
        print(f"❌ Erro na execução: {e}")

if __name__ == "__main__":
    exemplo_uso_completo()

